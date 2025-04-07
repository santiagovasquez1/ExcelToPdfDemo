import logging
import openpyxl

from src.domain.models.models import CellData, MergedRange, SheetData


class ExcelReader:
    def __init__(self, file_path):
        self.file_path = file_path
        self.wb = self._load_workbook()

    def _load_workbook(self):
        """
        Carga el archivo Excel y devuelve el objeto Workbook.
        """
        try:
            logging.info("Cargando el archivo Excel: {self.file_path}")
            return openpyxl.load_workbook(self.file_path, data_only=True)
        except FileNotFoundError as e:
            logging.error(f"El archivo {self.file_path} no fue encontrado.")
            raise e
        except openpyxl.utils.exceptions.InvalidFileException as e:
            logging.error(f"El archivo {self.file_path} no es un archivo Excel válido.")
            raise e
        except Exception as e:
            logging.error(f"Error al cargar el archivo Excel: {e}")
            raise e

    def read_excel(self) -> list[SheetData]:
        """
        Lee el archivo Excel y devuelve:
        - Una lista de objetos SheetData que representan cada hoja en el archivo.
        - Cada objeto SheetData contiene el nombre de la hoja, los datos de la tabla y los rangos fusionados.
        """

        all_data = []
        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]

            # Obtener los rangos fusionados y crearlos con Pydantic
            merged_ranges = []
            for merge in ws.merged_cells.ranges:
                merged_ranges.append(
                    MergedRange(
                        min_row=merge.min_row,
                        max_row=merge.max_row,
                        min_col=merge.min_col,
                        max_col=merge.max_col,
                    )
                )

            # Extraer los datos de cada celda y crear instancias de CellData
            table_data = []
            for row in ws.iter_rows():
                row_data = []
                for col_index, cell in enumerate(row, start=1):
                    row_data.append(
                        CellData(
                            value=cell.value,
                            row=cell.row,
                            col=col_index,  # Usamos el índice proporcionado por enumerate
                        )
                    )
                table_data.append(row_data)

            # Crear un objeto SheetData para la hoja actual
            sheet_data = SheetData(
                sheet_name=sheet_name,
                table_data=table_data,
                merged_ranges=merged_ranges,
            )
            all_data.append(sheet_data)
        return all_data
